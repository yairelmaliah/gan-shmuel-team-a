from db_utils import db_utils
from flask import request
import json
# import panda as pd

import os
from openpyxl import Workbook, load_workbook

mysql = db_utils()

def rates_handler():
  # file_name=request.args['file']
  file_name = os.path.basename("")
  wrkbk = load_workbook(f"in/rates.xlsx")
  sheet = wrkbk.active
  row = sheet.max_row
  column = sheet.max_column
  isheader = True
  for i in range(1, row + 1): 
    product_id = sheet.cell(row = i, column = 1) 
    scope = sheet.cell(row = i, column = 2) 
    if  i == 1:
      pass
    else:
      print(product_id.value)
     # data = mysql.getData(f'SELECT EXISTS(SELECT * FROM Rates WHERE product_id="{product_id}" AND scope = "{scope}");')
     # print(json.dumps(data)) 
     # print(i)
      


if __name__ == '__main__':
    rates_handler()



