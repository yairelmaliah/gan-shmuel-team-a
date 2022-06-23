from re import X
from flask import Flask, render_template, request
import json
from db_utils import db_utils
from openpyxl import Workbook, load_workbook
import os
import os.path

mysql = db_utils()



def POST_rates(file):
    
    file = str(file)
    file_name = f"/src/in/{file}"
    file_exists = os.path.exists(file_name)

    if file_exists == False:
        return (f"File : {file} does not exist, please try again." ,404)

    f = open("/src/in/last_file.txt", "w")
    f.write(file)
    f.close()


    wrkbk = load_workbook(f"{file_name}")
    sheet = wrkbk.active
    row = sheet.max_row
    column = sheet.max_column

    for i in range(1, row + 1): 
        product_id = sheet.cell(row = i, column = 1).value
        rate = sheet.cell(row = i, column = 2).value
        scope = sheet.cell(row = i, column = 3).value
        


        if  i != 1:
            data = mysql.getData(f'SELECT EXISTS(SELECT * FROM Rates WHERE product_id="{product_id}" AND scope = "{scope}");')
            dump = json.dumps(data)
            value_id = str(dump[-3])

            if value_id == str(0):

                mysql.setData(f'INSERT INTO Rates (`product_id`, `rate`, `scope`) VALUES ("{product_id}", {rate}, "{scope}");')
                                
            else:

                mysql.updateData(f'UPDATE `Rates` SET rate={rate} where product_id="{product_id}" and scope="{scope}";')

    new_rates =  mysql.getData(f'SELECT * FROM Rates;')
    r = json.dumps(new_rates)
    return r ,200
    
if __name__ == '__main__':
    POST_rates()
